import openpyxl

a = openpyxl.load_workbook('sample.xlsx')

#sheet = a.active
#for cellObj in sheet.ccolumns[1]:
#    print(cellObj.value)

values = a.get_sheet_by_name('Sheet1')
print('**********rows wise>>>>>>>>>>>>')
print(values['A2'].value)

b = values['B2']
print('**********rows wise>>>>>>>>>>>>')
print(str(b.value))

print('**********Cell Wise>>>>>>>>>>>>')
print(values['A2'].value + values['B2'].value + values['C2'].value)

c = values.cell(row=3, column=2).value
print('**********particular cell wise>>>>>>>>>>>>')
print('C : ', c)

print('**********Iterating over rows>>>>>>>>>>>>')
for i in range(1, 5, 1):
    print(i, ': ', values.cell(row =i, column=1).value, ' ' + values.cell(row =i, column=2).value, ' ' + values.cell(row =i, column=3).value )